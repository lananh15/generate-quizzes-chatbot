import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from common.base_app import QuizzSearchAppBase
from openai_handler import PineconeOpenAIHandler
from typing import List, Dict
import os
from openpyxl import Workbook, load_workbook
from flask import jsonify
from pinecone import Pinecone

class PineconeQuizzSearchApp(QuizzSearchAppBase, PineconeOpenAIHandler):
    def __init__(self, openai_handler, pinecone_index):
        super().__init__(openai_handler)
        self.pinecone_handler = pinecone_index

    # Lấy cấu trúc các chương, tiêu đề chính, tiêu đề phụ, tiểu mục
    def get_chapter_structure(self):
        query_embedding = self.openai_handler.openai.Embedding.create(
            input="chapter_title",
            model="text-embedding-ada-002"
        )['data'][0]['embedding']
        
        result = self.pinecone_handler.query(
            vector=query_embedding,
            top_k=100, 
            include_metadata=True
        )
        
        chapter_structure = {}
        for match in result['matches']:
            metadata = match['metadata']
            chapter_title = metadata.get('chapter_title')
            if chapter_title:
                if chapter_title not in chapter_structure:
                    chapter_structure[chapter_title] = {'headings': {}}
                
                heading_title = metadata.get('heading_title')
                if heading_title:
                    if heading_title not in chapter_structure[chapter_title]['headings']:
                        chapter_structure[chapter_title]['headings'][heading_title] = {'subheadings': {}}
                    
                    subheading_title = metadata.get('subheading_title')
                    if subheading_title:
                        if subheading_title not in chapter_structure[chapter_title]['headings'][heading_title]['subheadings']:
                            chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title] = []
                        
                        subsubheading_title = metadata.get('subsubheading_title')
                        if subsubheading_title:
                            chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title].append(subsubheading_title)
        
        return chapter_structure
    
    # Lưu kết quả truy vấn vào file excel
    def save_to_excel(self, keyword, results):
        base_filename = "search_results_pinecone_with_structured_data.xlsx"
        file_path = self.find_excel_file(base_filename)

        if file_path:
            print(f"Tìm thấy file tại: {file_path}")
        else:
            print(f"Không tìm thấy file {base_filename}. Sẽ tạo file mới trong thư mục hiện tại.")
            file_path = base_filename

        attempt = 1
        while True:
            try:
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Search Results"
                    ws.append(["Keyword", "Score", "Title", "Subtitle", "Content", "Highlights"])

                for result in results:
                    metadata = result['metadata']
                    title = metadata.get('chapter_title', '')
                    subtitle = ' > '.join(filter(None, [
                        metadata.get('heading_title', ''),
                        metadata.get('subheading_title', ''),
                        metadata.get('subsubheading_title', '')
                    ]))
                    
                    # Tìm nội dung từ các trường khác nhau
                    content = metadata.get('content', '')
                    if not content:
                        content = metadata.get('heading_content', '') or \
                                metadata.get('subheading_content', '') or \
                                metadata.get('subsubheading_content', '') or \
                                "Không có nội dung"
                    
                    highlights = ', '.join(metadata.get('keywords', []))

                    ws.append([
                        keyword,
                        result['score'],
                        title,
                        subtitle,
                        content,
                        highlights
                    ])

                wb.save(file_path)
                print(f"Kết quả đã được thêm vào file: {file_path}")
                break
            except PermissionError:
                print(f"Không thể ghi vào file {file_path}. Đang thử tạo file mới...")
                attempt += 1
                file_path = f"search_results_pinecone_with_structured_data_{attempt}.xlsx"
            except Exception as e:
                print(f"Lỗi khi lưu file: {e}")
                break

    # Truy vấn vào pinecone để tìm các vector tương đồng với keyword
    def search_pinecone(self, query: str, type: str) -> List[Dict]:
        number = {
            "chapter": 30,
            "heading": 20,
            "subheading": 15
        }.get(type, 5)

        query_embedding = self.openai_handler.openai.Embedding.create(
            input=query,
            model="text-embedding-ada-002"
        )['data'][0]['embedding']
            
        result = self.pinecone_handler.query(
            vector=query_embedding,
            top_k=number,
            include_metadata=True
        )
            
        sorted_results = sorted(result['matches'], key=lambda x: query.lower() in [k.lower() for k in x['metadata'].get('keywords', [])], reverse=True)
            
        print(f"Search results for query '{query}':")
        for i, match in enumerate(sorted_results, 1):
            print(f"Match {i}:")
            print(f"  Score: {match['score']}")
            print(f"  Metadata: {match['metadata']}")
            print("---")
            
        self.save_to_excel(query, sorted_results)
        return sorted_results

    # Hàm xử lý và response    
    def _handle_questions(self, message, question_type, max_questions):
        parts = message.split(':')
        if len(parts) != 3:
            return jsonify({"response": f"Vui lòng nhập theo định dạng: <code>'{question_type}: [tên {question_type}]: [số lượng câu hỏi (tối đa {max_questions})]'</code>."})

        keyword = parts[1].strip()

        try:
            num_questions = int(parts[2].strip())
            if not (0 < num_questions <= max_questions):
                return jsonify({"response": f"Số lượng câu hỏi phải là số dương và không quá {max_questions}."})

            pinecone_results = self.search_pinecone(keyword, question_type)

            # Tạo một hàm để kiểm tra từ khóa trong metadata
            def keyword_in_metadata(result, *keys):
                # Xử lý các giá trị từ metadata như danh sách hoặc chuỗi
                def get_value(key):
                    value = result['metadata'].get(key, '')
                    if isinstance(value, list):
                        # Chuyển danh sách thành chuỗi và kiểm tra từng phần tử
                        value = ' '.join(map(str, value))
                    return value.lower() if isinstance(value, str) else ''
                
                return any(keyword.lower() in get_value(key) for key in keys)

            # Lọc kết quả theo question_type
            if question_type == 'chapter':
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'chapter_title')]
            elif question_type == 'heading':
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'heading_title', 'keywords')]
            elif question_type == 'subheading':
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'subheading_title', 'heading_title', 'keywords')]
            else:  # 'subsubheading'
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'subsubheading_title', 'subheading_title', 'heading_title', 'keywords')]

            results_to_use = matched_results if matched_results else pinecone_results
            print(results_to_use)
            all_questions = self.openai_handler.generate_questions(results_to_use, num_questions)

            response = f"{num_questions} câu hỏi của {question_type} '{keyword}':\n\n"
            response += '\n----------\n'.join(f"{i+1}. {q}" for i, q in enumerate(all_questions))
            
            return jsonify({"response": response})
        except ValueError:
            return jsonify({"response": "Số lượng câu hỏi phải là một số nguyên."})
    
if __name__ == '__main__':
    openai_handler = PineconeOpenAIHandler("sk-proj-e3BgNgIvICywLYluJyeUT3BlbkFJenYTISe35HiZEiki9Gz3")
    pc = Pinecone(api_key="020a8257-5dd3-41f3-a710-53d7c6fac5d9")
    index = pc.Index("generate-quizz")
    app = PineconeQuizzSearchApp(openai_handler, index)
    app.run()