import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from common.base_app import QuizzSearchAppBase
from openai_handler import ElasticsearchOpenAIHandler
from elasticsearch import Elasticsearch
from typing import List, Dict, Tuple
import os
from openpyxl import Workbook, load_workbook
from flask import jsonify

class ElasticsearchQuizzSearchApp(QuizzSearchAppBase, ElasticsearchOpenAIHandler):
    def __init__(self, openai_handler, es_client):
        super().__init__(openai_handler)
        self.es_client = es_client
        self.index_name = 'qtda'

    # Lấy cấu trúc các chương, tiêu đề chính, tiêu đề phụ, tiểu mục
    def get_chapter_structure(self):
        query = {
            "size": 1000,
            "_source": ["title", "headings.title", "headings.subheadings.title", "headings.subheadings.subsubheadings.title"],
            "query": {
                "match_all": {}
            }
        }
        
        result = self.es_client.search(index=self.index_name, body=query)
        
        chapter_structure = {}
        for hit in result['hits']['hits']:
            source = hit['_source']
            chapter_title = source['title']
            if chapter_title not in chapter_structure:
                chapter_structure[chapter_title] = {'headings': {}}
            
            for heading in source.get('headings', []):
                heading_title = heading['title']
                if heading_title not in chapter_structure[chapter_title]['headings']:
                    chapter_structure[chapter_title]['headings'][heading_title] = {'subheadings': {}}
                
                for subheading in heading.get('subheadings', []):
                    subheading_title = subheading['title']
                    if subheading_title not in chapter_structure[chapter_title]['headings'][heading_title]['subheadings']:
                        chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title] = []
                    
                    for subsubheading in subheading.get('subsubheadings', []):
                        subsubheading_title = subsubheading['title']
                        if subsubheading_title not in chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title]:
                            chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title].append(subsubheading_title)
        
        return chapter_structure

    def _handle_chapter_structure(self):
        chapter_structure = self.get_chapter_structure()
        response = "Các chương được hỗ trợ:\n"
        for idx, (chapter_title, chapter_data) in enumerate(chapter_structure.items(), start=1):
            response += f"\n{idx}. {chapter_title}"
            for heading_title, heading_data in chapter_data['headings'].items():
                response += f"- {heading_title}"
                for subheading_title, subsubheadings in heading_data['subheadings'].items():
                    response += f"+ {subheading_title}"
                    for subsubheading in subsubheadings:
                        response += f"* {subsubheading}"
        response += "\n\n(trong đó số thứ tự là chương - là tiêu đề chính + là tiêu đề phụ * là tiểu mục)"
        response += "\n\nNhập <code>`chapter: [tên chương]: [số lượng câu hỏi (tối đa 25)]`</code> để tạo số lượng câu hỏi cho chương."
        response += "\n<code>`heading: [tên tiêu đề chính]: [số lượng câu hỏi (tối đa 15)]`</code> để tạo số lượng câu hỏi cho tiêu đề chính."
        response += "\n<code>`subheading: [tên tiêu đề phụ]: [số lượng câu hỏi (tối đa 10)]`</code> để tạo số lượng câu hỏi cho tiêu đề phụ."
        response += "\n<code>`subsubheading: [tên tiểu mục]: [số lượng câu hỏi (tối đa 5)]`</code> để tạo số lượng câu hỏi cho tiểu mục."

        response = response.replace("\n", "<br>")
        response = response.replace("\n", "<br>")
        response = response.replace("- ", "<br>&nbsp;&nbsp;- ")
        response = response.replace("+ ", "<br>&nbsp;&nbsp;&nbsp;&nbsp;+ ")
        response = response.replace("* ", "<br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;* ")    
        return jsonify({"response": response})

    # Lưu kết quả truy vấn vào file excel
    def save_search_result(self, keyword, search_results):
        base_filename = "search_results_elasticsearch.xlsx"
        file_path = self.find_excel_file(base_filename)

        if file_path:
            print(f"Tìm thấy file tại: {file_path}")
        else:
            print(f"Không tìm thấy file {base_filename}. Sẽ tạo file mới trong thư mục hiện tại.")
            file_path = base_filename

        attempt = 1
        while True:
            try:
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Search Results"
                    ws.append(["Keyword", "Score", "Title", "Subtitle", "Content", "Highlights"])

                for result in search_results:
                    highlights = "; ".join([f"{k}: {', '.join(v)}" for k, v in result['highlights'].items()])
                    ws.append([
                        keyword,
                        result['score'],
                        result['title'],
                        result['subtitle'],
                        result['content'],
                        highlights
                    ])

                wb.save(file_path)
                print(f"Kết quả đã được thêm vào file: {file_path}")
                break
            except PermissionError:
                print(f"Không thể ghi vào file {file_path}. Đang thử tạo file mới...")
                attempt += 1
                file_path = f"search_results_elasticsearch_{attempt}.xlsx"
            except Exception as e:
                print(f"Lỗi khi lưu file: {e}")
                break

    # Truy vấn vào elasticsearch để tìm các dữ liệu tương đồng với keyword
    def search(self, keyword: str, size: int = 10) -> List[Dict]:
        query = {
            "query": {
                "multi_match": {
                    "query": keyword,
                    "fields": ["*content", "*keywords", "*title"],
                    "type": "phrase_prefix"
                }
            },
            "highlight": {
                "fields": {
                    "*content": {},
                    "*keywords": {},
                    "*title": {}
                }
            }
        }
        
        result = self.es_client.search(index=self.index_name, body=query, size=size)
        
        search_results = []
        for hit in result['hits']['hits']:
            subtitle, content = self.find_content(hit['_source'], keyword)
            if not content:
                content = "Không tìm thấy nội dung cụ thể cho từ khóa này."
            
            hit_data = {
                "score": hit['_score'],
                "title": hit['_source'].get('title', ''),
                "subtitle": subtitle,
                "content": content,
                "highlights": hit.get('highlight', {})
            }
            search_results.append(hit_data)
            
        self.save_search_result(keyword, search_results)
        print(search_results)
        return search_results
    
    # Hàm xử lý và response
    def _handle_questions(self, message, question_type, max_questions):
        parts = message.split(':')
        if len(parts) != 3:
            return jsonify({"response": f"Vui lòng nhập theo định dạng: <code>'{question_type}: [tên {question_type}]: [số lượng câu hỏi (tối đa {max_questions})]'</code>."})
        
        keyword = parts[1].strip()
        try:
            num_questions = int(parts[2].strip())
            if num_questions <= 0 or num_questions > max_questions:
                return jsonify({"response": f"Số lượng câu hỏi phải là số dương và không quá {max_questions}."})
            
            search_results = self.search(keyword)
            if not search_results:
                return jsonify({"response": f"Không tìm thấy nội dung cho {question_type} '{keyword}'."})
            
            content = search_results[0]['content']
            all_questions = self.openai_handler.generate_questions(content, num_questions)
            
            response = f"{num_questions} câu hỏi của {question_type} '{keyword}':\n\n"
            for i, question in enumerate(all_questions, 1):
                response += f"{i}. {question}\n----------\n"
            return jsonify({"response": response})
        except ValueError:
            return jsonify

    # Thu thập nội dung từ đối tượng có cấu trúc lồng nhau, bao gồm tiêu đề, nội dung...    
    def collect_content(self, obj):
        contents = []
        if 'title' in obj:
            contents.append(f"{obj['title']}:")
        if 'content' in obj:
            contents.append(f"{obj['content']}\n")
        if 'headings' in obj:
            for heading in obj['headings']:
                contents.append(self.collect_content(heading))
        if 'subheadings' in obj:
            for subheading in obj['subheadings']:
                contents.append(self.collect_content(subheading))
        if 'subsubheadings' in obj:
            for subsubheading in obj['subsubheadings']:
                contents.append(self.collect_content(subsubheading))
        return "\n".join(contents)

    # Tìm kiếm nội dung trong cấu trúc dữ liệu lồng nhau dựa trên từ khóa và trả về đường dẫn và nội dung của đối tượng chứa từ khóa đó
    def find_content(self, source: Dict, keyword: str) -> Tuple[str, str]:
        def search_nested(obj, keyword, path=[]):
            if isinstance(obj, dict):
                title = obj.get('title', '').lower()
                content = obj.get('content', '').lower()
                keywords = ' '.join(obj.get('keywords', [])).lower()
                
                if keyword.lower() in title or keyword.lower() in content or keyword.lower() in keywords:
                    current_path = ' > '.join(path + [obj.get('title', '')])
                    current_content = self.collect_content(obj)
                    return current_path, current_content
                
                for key, value in obj.items():
                    if key in ['headings', 'subheadings', 'subsubheadings']:
                        result = search_nested(value, keyword, path + [obj.get('title', '')])
                        if result[0] or result[1]:
                            return result
            elif isinstance(obj, list):
                for item in obj:
                    result = search_nested(item, keyword, path)
                    if result[0] or result[1]:
                        return result
            return None, None

        return search_nested(source, keyword)
    
if __name__ == '__main__':
    openai_handler = ElasticsearchOpenAIHandler("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")
    es_client = Elasticsearch(['http://localhost:9200'])
    app = ElasticsearchQuizzSearchApp(openai_handler, es_client)
    app.run()