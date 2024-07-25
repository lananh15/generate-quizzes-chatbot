import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from common.base_app import QuizzSearchAppBase
from common.handlers.llm_handler_base import LLMHandlerBase
from content_processor import ContentProcessor
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from flask import jsonify
from pinecone import Pinecone
from openai import OpenAI

class PineconeQuizzSearchApp(QuizzSearchAppBase, LLMHandlerBase):
    def __init__(self, llm_handler, pinecone_index, openai_client):
        super().__init__(llm_handler)
        self.pinecone_handler = pinecone_index
        self.openai_client = openai_client

    # Lấy cấu trúc các chương, tiêu đề chính, tiêu đề phụ, tiểu mục
    def get_chapter_structure(self):
        response = self.openai_client.embeddings.create(
            input="chapter_title",
            model="text-embedding-ada-002"  # Hoặc model phù hợp khác
        )
        query_embedding = response.data[0].embedding

        # Truy vấn Pinecone
        result = self.pinecone_handler.query(
            vector=query_embedding,
            top_k=100,
            include_metadata=True
        )
        
        chapter_structure = {}
        for match in result['matches']:
            metadata = match['metadata']
            chapter_title = metadata.get('chapter_title')
            if chapter_title:
                if chapter_title not in chapter_structure:
                    chapter_structure[chapter_title] = {'headings': {}}
                
                heading_title = metadata.get('heading_title')
                if heading_title:
                    if heading_title not in chapter_structure[chapter_title]['headings']:
                        chapter_structure[chapter_title]['headings'][heading_title] = {'subheadings': {}}
                    
                    subheading_title = metadata.get('subheading_title')
                    if subheading_title:
                        if subheading_title not in chapter_structure[chapter_title]['headings'][heading_title]['subheadings']:
                            chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title] = []
                        
                        subsubheading_title = metadata.get('subsubheading_title')
                        if subsubheading_title:
                            chapter_structure[chapter_title]['headings'][heading_title]['subheadings'][subheading_title].append(subsubheading_title)
        
        return chapter_structure
    
    def _handle_chapter_structure(self):
        chapter_structure = self.get_chapter_structure()
        response = "Các chương được hỗ trợ:\n"
        for idx, (chapter_title, chapter_data) in enumerate(chapter_structure.items(), start=1):
            response += f"\n{idx}. {chapter_title}"
            for heading_title, heading_data in chapter_data['headings'].items():
                response += f"- {heading_title}"
                for subheading_title, subsubheadings in heading_data['subheadings'].items():
                    response += f"+ {subheading_title}"
                    for subsubheading in subsubheadings:
                        response += f"* {subsubheading}"
        response += "\n\n(trong đó số thứ tự là chương - là tiêu đề chính + là tiêu đề phụ * là tiểu mục)"
        response += "\n\nNhập <code>`chapter: [tên chương]: [số lượng câu hỏi (tối đa 25)]`</code> để tạo số lượng câu hỏi cho chương."
        response += "\n<code>`heading: [tên tiêu đề chính]: [số lượng câu hỏi (tối đa 15)]`</code> để tạo số lượng câu hỏi cho tiêu đề chính."
        response += "\n<code>`subheading: [tên tiêu đề phụ]: [số lượng câu hỏi (tối đa 10)]`</code> để tạo số lượng câu hỏi cho tiêu đề phụ."
        response += "\n<code>`subsubheading: [tên tiểu mục]: [số lượng câu hỏi (tối đa 5)]`</code> để tạo số lượng câu hỏi cho tiểu mục."

        response += "\n\nNhập <code>`mode: openai`</code> hoặc <code>`mode: gemini`</code> để chọn chế độ và <code>`mode`</code> để xem chế độ hiện tại."
        response = response.replace("\n", "<br>")
        response = response.replace("\n", "<br>")
        response = response.replace("- ", "<br>&nbsp;&nbsp;- ")
        response = response.replace("+ ", "<br>&nbsp;&nbsp;&nbsp;&nbsp;+ ")
        response = response.replace("* ", "<br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;* ")    
        return jsonify({"response": response})

    # Lưu kết quả truy vấn vào file excel
    def save_to_excel(self, keyword, results):
        base_filename = "search_results_pinecone_with_structured_data.xlsx"
        file_path = self.find_excel_file(base_filename)

        if file_path:
            print(f"Tìm thấy file tại: {file_path}")
        else:
            print(f"Không tìm thấy file {base_filename}. Sẽ tạo file mới trong thư mục hiện tại.")
            file_path = base_filename

        attempt = 1
        while True:
            try:
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Search Results"
                    ws.append(["Keyword", "Score", "Title", "Subtitle", "Content", "Highlights"])

                for result in results:
                    metadata = result['metadata']
                    title = metadata.get('chapter_title', '')
                    subtitle = ' > '.join(filter(None, [
                        metadata.get('heading_title', ''),
                        metadata.get('subheading_title', ''),
                        metadata.get('subsubheading_title', '')
                    ]))
                    
                    # Tìm nội dung từ các trường khác nhau
                    content = metadata.get('content', '')
                    if not content:
                        content = metadata.get('heading_content', '') or \
                                metadata.get('subheading_content', '') or \
                                metadata.get('subsubheading_content', '') or \
                                "Không có nội dung"
                    
                    highlights = ', '.join(metadata.get('keywords', []))

                    ws.append([
                        keyword,
                        result['score'],
                        title,
                        subtitle,
                        content,
                        highlights
                    ])

                wb.save(file_path)
                print(f"Kết quả đã được thêm vào file: {file_path}")
                break
            except PermissionError:
                print(f"Không thể ghi vào file {file_path}. Đang thử tạo file mới...")
                attempt += 1
                file_path = f"search_results_pinecone_with_structured_data_{attempt}.xlsx"
            except Exception as e:
                print(f"Lỗi khi lưu file: {e}")
                break

    # Truy vấn vào pinecone để tìm các vector tương đồng với keyword
    def search_pinecone(self, query: str, type: str) -> List[Dict]:
        number = {
            "chapter": 30,
            "heading": 20,
            "subheading": 15
        }.get(type, 5)

        response = self.openai_client.embeddings.create(
            input=query,
            model="text-embedding-ada-002"  # Hoặc model phù hợp khác
        )
        query_embedding = response.data[0].embedding

        # Truy vấn Pinecone
        result = self.pinecone_handler.query(
            vector=query_embedding,
            top_k=number,
            include_metadata=True
        )
            
        sorted_results = sorted(result['matches'], key=lambda x: query.lower() in [k.lower() for k in x['metadata'].get('keywords', [])], reverse=True)
            
        print(f"Search results for query '{query}':")
        for i, match in enumerate(sorted_results, 1):
            print(f"Match {i}:")
            print(f"  Score: {match['score']}")
            print(f"  Metadata: {match['metadata']}")
            print("---")
            
        self.save_to_excel(query, sorted_results)
        return sorted_results

    # Hàm xử lý và response    
    def _handle_questions(self, message, question_type, max_questions):
        parts = message.split(':')
        if len(parts) != 3:
            return jsonify({"response": f"Vui lòng nhập theo định dạng: <code>'{question_type}: [tên {question_type}]: [số lượng câu hỏi (tối đa {max_questions})]'</code>."})

        keyword = parts[1].strip()

        try:
            num_questions = int(parts[2].strip())
            if not (0 < num_questions <= max_questions):
                return jsonify({"response": f"Số lượng câu hỏi phải là số dương và không quá {max_questions}."})

            pinecone_results = self.search_pinecone(keyword, question_type)
            print(pinecone_results)

            # Tạo một hàm để kiểm tra từ khóa trong metadata
            def keyword_in_metadata(result, *keys):
                # Xử lý các giá trị từ metadata như danh sách hoặc chuỗi
                def get_value(key):
                    value = result['metadata'].get(key, '')
                    if isinstance(value, list):
                        # Chuyển danh sách thành chuỗi và kiểm tra từng phần tử
                        value = ' '.join(map(str, value))
                    return value.lower() if isinstance(value, str) else ''
                
                return any(keyword.lower() in get_value(key) for key in keys)

            # Lọc kết quả theo question_type
            if question_type == 'chapter':
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'chapter_title')]
            elif question_type == 'heading':
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'heading_title', 'keywords')]
            elif question_type == 'subheading':
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'subheading_title', 'heading_title', 'keywords')]
            else:  # 'subsubheading'
                matched_results = [result for result in pinecone_results if keyword_in_metadata(result, 'subsubheading_title', 'subheading_title', 'heading_title', 'keywords')]

            results_to_use = matched_results if matched_results else pinecone_results
            print(results_to_use)
            all_questions = self.llm_handler.generate_questions(results_to_use, num_questions)

            response = f"{num_questions} câu hỏi của {question_type} '{keyword}':\n\n"
            response += '\n----------\n'.join(f"{i+1}. {q}" for i, q in enumerate(all_questions))
            
            return jsonify({"response": response})
        except ValueError:
            return jsonify({"response": "Số lượng câu hỏi phải là một số nguyên."})
    
if __name__ == '__main__':
    OPENAI_API_KEY = "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"
    GOOGLE_API_KEY = "AIzaSyAYxPv1wiS66B0qjiTO59R6t1V5j27dcrY"
    PINECONE_API_KEY = "020a8257-5dd3-41f3-a710-53d7c6fac5d9"
    ANTHROPIC_API_KEY = "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"

    llm_handler = ContentProcessor(OPENAI_API_KEY, GOOGLE_API_KEY, ANTHROPIC_API_KEY)
    pc = Pinecone(api_key=PINECONE_API_KEY)
    index = pc.Index("generate-quizz")

    openai_client = OpenAI(api_key=OPENAI_API_KEY)

    app = PineconeQuizzSearchApp(llm_handler, index, openai_client)
    app.run()