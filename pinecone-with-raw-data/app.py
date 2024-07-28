import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from common.base_app import QuizzSearchAppBase
from common.handlers.llm_handler_base import LLMHandlerBase
from content_processor import ContentProcessor
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from flask import jsonify
from pinecone import Pinecone
from openai import OpenAI

class PineconeQuizzSearchApp(QuizzSearchAppBase, LLMHandlerBase):
    def __init__(self, llm_handler, pinecone_index, openai_client):
        super().__init__(llm_handler)
        self.pinecone_handler = pinecone_index
        self.openai_client = openai_client
    
    def _handle_chapter_structure(self):
        response = "1. Tổng quan về quản lý dự án\n2. Cơ cấu quản lý dự án\n3. Quy trình quản lý dự án\n4. Quản lý phạm vi\n5. Quản lý thời gian\n6. Quản lý chi phí\n7. Quản lý rủi ro\n8. Quản lý chất lượng"
        response += "\n\nNhập <code>`keyword: [keyword]: [số lượng câu hỏi (tối đa 20)]`</code> để tạo số lượng câu hỏi cho keyword bạn cần."
        return jsonify({"response": response})
    
    # Lưu kết quả truy vấn vào file excel
    def save_to_excel(self, keyword: str, results: List[Dict]):
        base_filename = "search_results_pinecone_with_raw_data.xlsx"
        file_path = self.find_excel_file(base_filename)

        if file_path:
            print(f"Tìm thấy file tại: {file_path}")
        else:
            print(f"Không tìm thấy file {base_filename}. Sẽ tạo file mới trong thư mục hiện tại.")
            file_path = base_filename

        attempt = 1
        while True:
            try:
                if os.path.exists(file_path):
                    wb = load_workbook(file_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Search Results"
                    ws.append(["Keyword", "Score", "Content"])  # Cột tiêu đề chỉ bao gồm: Keyword, Score, Content

                for result in results:
                    metadata = result['metadata']
                    content = metadata.get('text', 'Không có nội dung')  # Dữ liệu hiện có chỉ có trường 'text'

                    ws.append([
                        keyword,
                        result['score'],
                        content
                    ])

                wb.save(file_path)
                print(f"Kết quả đã được thêm vào file: {file_path}")
                break
            except PermissionError:
                print(f"Không thể ghi vào file {file_path}. Đang thử tạo file mới...")
                attempt += 1
                file_path = f"search_results_pinecone_with_raw_data_{attempt}.xlsx"
            except Exception as e:
                print(f"Lỗi khi lưu file: {e}")
                break

    # Truy vấn vào pinecone để tìm các vector tương đồng với keyword
    def search_pinecone(self, query: str) -> List[Dict]:
        # Tạo embedding cho truy vấn
        response = self.openai_client.embeddings.create(
            input=query,
            model="text-embedding-ada-002"
        )
        query_embedding = response.data[0].embedding

        # Truy vấn Pinecone
        result = self.pinecone_handler.query(
            vector=query_embedding,
            top_k=4,
            include_metadata=True
        )

        # Sắp xếp kết quả theo điểm số, giữ nguyên metadata
        sorted_results = sorted(result['matches'], key=lambda x: x['score'], reverse=True)
        # Lưu kết quả vào file Excel
        self.save_to_excel(query, sorted_results)

        return sorted_results

    # Hàm xử lý và response    
    def _handle_questions_with_keyword(self, message, max_questions):
        parts = message.split(':')
        if len(parts) != 3:
            return jsonify({"response": f"Vui lòng nhập theo định dạng: <code>'keyword: [tên keyword]: [số lượng câu hỏi (tối đa {max_questions})]'</code>."})

        keyword = parts[1].strip()

        try:
            num_questions = int(parts[2].strip())
            if not (0 < num_questions <= max_questions):
                return jsonify({"response": f"Số lượng câu hỏi phải là số dương và không quá {max_questions}."})

            pinecone_results = self.search_pinecone(keyword)
            print(pinecone_results)
            keyword_parts = keyword.lower().split()
            results_to_use = [
                result for result in pinecone_results 
                if all(part in result['metadata']['text'].lower() for part in keyword_parts)
            ]
            print(f"\nCác kết quả dùng để sinh quizz:\n")
            print(results_to_use)

            all_questions = self.llm_handler.generate_questions(results_to_use, num_questions)

            response = f"{num_questions} câu hỏi của '{keyword}':\n\n"
            response += '\n----------\n'.join(f"{i+1}. {q}" for i, q in enumerate(all_questions))
            
            return jsonify({"response": response})
        except ValueError:
            return jsonify({"response": "Số lượng câu hỏi phải là một số nguyên."})
    
if __name__ == '__main__':
    OPENAI_API_KEY = "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"
    GOOGLE_API_KEY = "AIzaSyAYxPv1wiS66B0qjiTO59R6t1V5j27dcrY"
    PINECONE_API_KEY = "020a8257-5dd3-41f3-a710-53d7c6fac5d9"
    ANTHROPIC_API_KEY = "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"

    llm_handler = ContentProcessor(OPENAI_API_KEY, GOOGLE_API_KEY, ANTHROPIC_API_KEY)
    pc = Pinecone(api_key=PINECONE_API_KEY)
    index = pc.Index("generate-quizz-with-raw-data")

    openai_client = OpenAI(api_key=OPENAI_API_KEY)

    app = PineconeQuizzSearchApp(llm_handler, index, openai_client)
    app.run()